import os
import argparse
from datetime import datetime

import pandas as pd

def compare_cutting_files(file1, file2, output_file):
    """
    比較兩個 Excel 檔案中相同 sub_program 和 row_id 的列，並輸出比較結果。
    
    參數:
    file1 (str): 第一個 Excel 檔案的路徑
    file2 (str): 第二個 Excel 檔案的路徑
    output_file (str): 輸出結果的檔案路徑
    """
    print(f"正在讀取檔案 {file1}...")
    df1 = pd.read_excel(file1)
    print(f"正在讀取檔案 {file2}...")
    df2 = pd.read_excel(file2)
    
    # 確保必要欄位存在
    required_columns = ["sub_program", "row_id", "move_code"]
    for df, file in [(df1, file1), (df2, file2)]:
        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            raise ValueError(f"檔案 {file} 缺少必要欄位: {', '.join(missing_cols)}")
    
    # 篩選 move_code 屬於特定類型的資料列
    valid_move_codes = ["G01", "G02", "G03", "G81", "G82", "G83"]

    # 使用 .copy() 創建明確的副本，避免 SettingWithCopyWarning
    df1 = df1[df1["move_code"].isin(valid_move_codes)].copy()
    df2 = df2[df2["move_code"].isin(valid_move_codes)].copy()
    
    print(f"檔案1篩選後: {len(df1)}列")
    print(f"檔案2篩選後: {len(df2)}列")
    
    # 將 sub_program 和 row_id 轉為字串，確保比對一致性
    df1['sub_program'] = df1['sub_program'].astype(str)
    df2['sub_program'] = df2['sub_program'].astype(str)
    df1['row_id'] = df1['row_id'].astype(str)
    df2['row_id'] = df2['row_id'].astype(str)
    
    # 建立唯一識別碼
    df1['unique_id'] = df1['sub_program'] + "_" + df1['row_id']
    df2['unique_id'] = df2['unique_id'] = df2['sub_program'] + "_" + df2['row_id']
    
    # 找出要比較的所有欄位
    compare_columns = [
        "cutting_area", "hit_area", "ap_sum_voxel", "path_area_xy", 
        "hit_area_xy", "ae_sum_voxel", "path_area_z", "hit_area_z"
    ]
    
    # 確保比較欄位在兩個 DataFrame 中都存在
    for df, file in [(df1, file1), (df2, file2)]:
        missing_cols = [col for col in compare_columns if col not in df.columns]
        if missing_cols:
            print(f"警告: 檔案 {file} 缺少以下比較欄位: {', '.join(missing_cols)}")
            # 移除不存在的欄位
            compare_columns = [col for col in compare_columns if col not in missing_cols]
    
    if not compare_columns:
        raise ValueError("沒有可比較的欄位，請檢查輸入檔案")
    
    print(f"將比較以下欄位: {', '.join(compare_columns)}")
    
    # 找出兩個檔案共同的 unique_id
    common_ids = set(df1['unique_id']) & set(df2['unique_id'])
    print(f"找到 {len(common_ids)} 個相同 sub_program 和 row_id 的列")
    
    if not common_ids:
        raise ValueError("沒有找到相同 sub_program 和 row_id 的列，無法進行比較")
    
    # 準備結果 DataFrame
    results = []
    
    # 對每個共同的 ID 進行比較
    for unique_id in common_ids:
        row1 = df1[df1['unique_id'] == unique_id].iloc[0]
        row2 = df2[df2['unique_id'] == unique_id].iloc[0]
        
        sub_program = row1['sub_program']
        row_id = row1['row_id']
        
        result_row = {
            'sub_program': sub_program,
            'row_id': row_id
        }
        
        # 比較每個欄位的值
        for col in compare_columns:
            val1 = row1[col]
            val2 = row2[col]
            
            # 檢查值是否匹配
            if pd.isna(val1) or pd.isna(val2):
                # 如果任何一個值是 NaN，則不匹配
                match = False
            else:
                # 嘗試將兩個值都轉為整數進行比較
                try:
                    # 嚴格將兩個值轉換為整數並比較
                    int_val1 = int(val1)
                    int_val2 = int(val2)
                    match = (int_val1 == int_val2)
                except (ValueError, TypeError):
                    # 如果無法轉為整數，視為不匹配
                    match = False
            
            # 將比較結果和兩個檔案的值加入結果 DataFrame
            result_row[f'{col}_file1'] = val1
            result_row[f'{col}_file2'] = val2
            result_row[f'{col}_diff'] = not match
            
        results.append(result_row)
    
    # 將結果轉換為 DataFrame
    results_df = pd.DataFrame(results)
    
    # 計算每列的差異欄位數量
    diff_columns = [col for col in results_df.columns if col.endswith('_diff')]
    results_df['diff_count'] = results_df[diff_columns].sum(axis=1)
    
    # 總共有差異的列數
    rows_with_diff = (results_df['diff_count'] > 0).sum()
    print(f"共有 {rows_with_diff} 列存在差異")

    # 將 row_id 轉換為數值型別以便正確排序
    # 首先備份原始的 row_id
    results_df['original_row_id'] = results_df['row_id']

    # 嘗試將 row_id 轉換為數值型別
    results_df['row_id_numeric'] = pd.to_numeric(results_df['row_id'], errors='coerce')

    # 使用數值型別 row_id 進行排序
    results_df = results_df.sort_values(by=['sub_program', 'row_id_numeric'])

    # 恢復原始的 row_id 並刪除輔助欄位
    results_df['row_id'] = results_df['original_row_id']
    results_df.drop(['original_row_id', 'row_id_numeric'], axis=1, inplace=True)
    
    # 在儲存結果到 Excel 部分進行修改
    # 儲存結果到 Excel
    print(f"正在保存結果至 {output_file}...")
    
    # 使用 ExcelWriter 並指定 engine 為 openpyxl
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False, sheet_name='比較結果')
        
        # 獲取 worksheet
        worksheet = writer.sheets['比較結果']
        
        # 設定黃色底色格式
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # 找出有差異的列的索引
        rows_with_diff_indices = [i + 2 for i, diff in enumerate(results_df['diff_count'] > 0) if diff]
        # 注意：+2 是因為 Excel 從 1 開始計數，且第 1 列是標題
        
        # 對每個有差異的列套用黃色底色
        for row_idx in rows_with_diff_indices:
            for col_idx in range(1, len(results_df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # 保存原始值
                original_value = cell.value
                # 設置黃色底色
                cell.fill = yellow_fill
                # 確保值不變
                cell.value = original_value
        
    print(f"結果已保存至 {output_file}")

def main():
    parser = argparse.ArgumentParser(description='比較兩個 CNC Simulation Cutting 檔案中數據的差異')
    parser.add_argument('file1', help='第一個 Cutting 檔案的路徑')
    parser.add_argument('file2', help='第二個 Cutting 檔案的路徑')
    parser.add_argument('-o', '--output', help='輸出結果的 Excel 檔案路徑',
                       default=None)
    
    args = parser.parse_args()
    
    # 檢查輸入檔案是否存在
    for file in [args.file1, args.file2]:
        if not os.path.exists(file):
            print(f"錯誤: 找不到檔案 {file}")
            return
    
    # 從檔案路徑中提取檔案名稱（不含副檔名）
    file1_name = os.path.splitext(os.path.basename(args.file1))[0]
    file2_name = os.path.splitext(os.path.basename(args.file2))[0]
    
    # 如果沒有指定輸出檔案，則使用來源檔案名稱來命名
    if args.output is None:
        output_file = f'Cutting_比較_{file1_name}_vs_{file2_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:
        output_file = args.output

    try:
        compare_cutting_files(args.file1, args.file2, output_file)
        
    except Exception as e:
        print(f"比較過程中發生錯誤: {e}")

if __name__ == "__main__":
    main()